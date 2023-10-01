from openpyxl.reader.excel import load_workbook


class Excel:

    def __init__(self, filename):
        self.filename = filename
        self.workbook = load_workbook(filename)
        self.sheet = self.workbook.active

    def set_sheet(self, sheet_name):
        self.sheet = self.workbook[sheet_name]

    def get_sheet_names(self):
        return self.workbook.sheetnames

    def get_rows_as_list(self):
        data = []
        for i, row in enumerate(self.sheet):
            # Get keys from first row
            if i == 0:
                keys = [cell.value for cell in row]
                continue
            data.append({keys[j]: cell.value for j, cell in enumerate(row)})
        return data

    def get_cols_as_list(self, skip_col=None):
        skip_col = skip_col or []
        data = []
        for i, col in enumerate(self.sheet.iter_cols()):
            # Get keys from first col
            if i in skip_col:
                continue
            if i == 0:
                keys = [cell.value or "header" for cell in col]
                continue
            data.append({keys[j]: cell.value for j, cell in enumerate(col)})
        return data

    def add_data_to_table(self, data, additional_data=None):
        additional_data = additional_data or {}
        row_num = 0
        for i, row in enumerate(self.sheet):
            row_num = i + 1
            if i == 0:
                keys = [cell.value for cell in row]
                continue
        row_num += 1
        for item in data:
            for j, key in enumerate(keys):
                if key == "id":
                    value = row_num - 1
                else:
                    value = additional_data.get(key) or item.get(key)
                self.sheet.cell(row_num, j + 1, value)
            row_num += 1

    def save(self, filename):
        self.workbook.save(filename)
